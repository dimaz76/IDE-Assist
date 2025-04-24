import os
import sys
import json
import yaml
import csv
import logging
import argparse
import pandas as pd
from openpyxl import Workbook

# Патчим pandas.read_excel, чтобы по умолчанию читать все значения как строки
_original_read_excel = pd.read_excel
pd.read_excel = lambda io, *args, **kwargs: _original_read_excel(io, *args, dtype=str, **kwargs)

logging.basicConfig(level=logging.INFO)

def load_config(path: str) -> tuple[str, str, int]:
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Config file '{path}' not found.")
    ext = os.path.splitext(path)[1].lower()
    with open(path, 'r', encoding='utf-8') as f:
        if ext in (".yaml", ".yml"):
            cfg = yaml.safe_load(f)
        elif ext == ".json":
            cfg = json.load(f)
        else:
            raise ValueError(f"Unsupported config format '{ext}'")
    for key in ("input", "output"):
        if key not in cfg:
            raise KeyError(f"Config file is missing key: {key}")
    raw_thr = cfg.get("threshold", 10)
    try:
        thr = int(raw_thr)
    except Exception:
        raise ValueError(f"Invalid threshold '{raw_thr}'")
    return cfg["input"], cfg["output"], thr

def load_data(path: str) -> list[dict]:
    if not os.path.isfile(path):
        raise FileNotFoundError(f"Data file '{path}' not found.")
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        df = pd.read_csv(path, dtype=str)
    elif ext == ".json":
        try:
            df = pd.read_json(path, dtype=str)
        except ValueError:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            df = pd.DataFrame(data)
    elif ext in (".xls", ".xlsx"):
        df = pd.read_excel(path)  # patched to dtype=str
    else:
        raise ValueError(f"Unsupported data format '{ext}'")
    return df.to_dict(orient="records")

def process_data(data: list[dict], threshold: int = 10) -> list[dict]:
    result = []
    for row in data:
        try:
            if int(row.get("value", 0)) > threshold:
                result.append({"id": row.get("id", ""), "value": row.get("value", "")})
        except Exception:
            continue
    return result

def save_result(data: list[dict], path: str, dry_run: bool = False) -> None:
    if dry_run:
        logging.info("Dry run: would save %d rows to %s", len(data), path)
        return
    if not data:
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(["id", "value"])
        return
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=list(data[0].keys()))
            writer.writeheader()
            writer.writerows(data)
    elif ext == ".json":
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    elif ext in (".xls", ".xlsx"):
        wb = Workbook()
        ws = wb.active
        headers = list(data[0].keys())
        # Запись заголовка
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.data_type = 's'
        # Запись данных как строки
        for row_idx, row in enumerate(data, start=2):
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=str(row[h]))
                cell.data_type = 's'
        wb.save(path)
    else:
        raise ValueError(f"Unsupported output format '{ext}'")

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(prog="main.py", description="IDE-Assist script skeleton")
    parser.add_argument("-c", "--config", default="config.yaml", help="Path to config file")
    parser.add_argument("--dry-run", action="store_true", help="Do not write output")
    parser.add_argument("-v", "--verbose", action="store_true", help="Verbose logging")
    return parser.parse_args()

def main() -> None:
    args = parse_args()
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    inp, outp, thr = load_config(args.config)
    data = load_data(inp)
    processed = process_data(data, threshold=thr)
    save_result(processed, outp, dry_run=args.dry_run)

if __name__ == "__main__":
    main()
